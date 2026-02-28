import os
import calendar
import datetime as dt
from typing import Dict, Any, List, Optional, Tuple

import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
APP_TITLE = "Life Tracker Dashboard"
EXCEL_FILE = "life_tracker.xlsx"

DAILY_SHEET = "daily"
PROJECTS_SHEET = "projects"
MONTHLY_SHEET = "monthly"

DAILY_HEADERS = [
    "date", "income", "spent",
    "dsa_leetcode_solved",
    "js_fe_min", "js_be_min",
    "prayer_fajr", "prayer_dhuhr", "prayer_asr", "prayer_maghrib", "prayer_isha",
    "eng_listen_min", "eng_read_min", "eng_speak_min", "eng_write_min",
    "gym_trained", "calories", "screen_time_min",
    "notes"
]

PROJECT_HEADERS = [
    "date", "month", "track", "project_name",
    "project_level", "hours_spent", "link", "notes"
]

MONTHLY_HEADERS = [
    "month",
    "eng_listen_level", "eng_read_level", "eng_speak_level", "eng_write_level",
    "dsa_level", "js_fe_level", "js_be_level",
    "body_fat_pct",
    "notes"
]


# =========================
# THEME / STYLE
# =========================
DARK_CSS = """
<style>
/* App background */
.stApp {
  background-color: #0b0f17;
  color: #e6edf3;
}

/* Sidebar background */
section[data-testid="stSidebar"] {
  background: #0a0d14;
  border-right: 1px solid rgba(255,255,255,0.06);
}

/* Headings */
h1, h2, h3, h4, h5, h6 { color: #e6edf3; }

/* Make default widgets look nicer */
div[data-baseweb="input"] > div,
div[data-baseweb="textarea"] > div,
div[data-baseweb="select"] > div {
  background-color: #111827 !important;
  border: 1px solid rgba(255,255,255,0.10) !important;
}

.stButton button {
  background: linear-gradient(90deg, #2563eb, #7c3aed);
  border: 0;
  color: white;
  font-weight: 600;
  padding: 0.6rem 1rem;
  border-radius: 12px;
}

.stButton button:hover {
  filter: brightness(1.05);
}

/* Card */
.card {
  background: #0f172a;
  border: 1px solid rgba(255,255,255,0.08);
  border-radius: 16px;
  padding: 14px 16px;
  box-shadow: 0 8px 24px rgba(0,0,0,0.35);
}
.card .label { opacity: 0.75; font-size: 0.85rem; }
.card .value { font-size: 1.6rem; font-weight: 700; margin-top: 4px; }
.card .sub { opacity: 0.75; margin-top: 2px; font-size: 0.85rem; }

/* Divider */
.hr {
  height: 1px;
  background: rgba(255,255,255,0.08);
  margin: 14px 0;
}

/* Tables */
[data-testid="stDataFrame"] {
  border: 1px solid rgba(255,255,255,0.08);
  border-radius: 14px;
  overflow: hidden;
}
</style>
"""


# =========================
# UTIL
# =========================
def today_date() -> dt.date:
    return dt.date.today()

def month_str(d: dt.date) -> str:
    return f"{d.year:04d}-{d.month:02d}"

def parse_month(s: str) -> Tuple[int, int]:
    y, m = s.split("-")
    y, m = int(y), int(m)
    if not (1 <= m <= 12):
        raise ValueError("Month must be YYYY-MM")
    return y, m

def month_date_range(year: int, month: int) -> Tuple[dt.date, dt.date]:
    last = calendar.monthrange(year, month)[1]
    return dt.date(year, month, 1), dt.date(year, month, last)

def safe_float(x, default=0.0) -> float:
    try:
        if x is None or x == "":
            return float(default)
        return float(x)
    except Exception:
        return float(default)

def safe_int(x, default=0) -> int:
    try:
        if x is None or x == "":
            return int(default)
        return int(float(x))
    except Exception:
        return int(default)

def prayer_count_row(row: pd.Series) -> int:
    return int(safe_int(row.get("prayer_fajr"))) + int(safe_int(row.get("prayer_dhuhr"))) + int(safe_int(row.get("prayer_asr"))) + int(safe_int(row.get("prayer_maghrib"))) + int(safe_int(row.get("prayer_isha")))

def ensure_plotly_dark(fig):
    fig.update_layout(template="plotly_dark", margin=dict(l=10, r=10, t=40, b=10))
    return fig

def card(label: str, value: str, sub: str = ""):
    st.markdown(
        f"""
        <div class="card">
          <div class="label">{label}</div>
          <div class="value">{value}</div>
          <div class="sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

def info_box(title: str, text: str):
    st.markdown(
        f"""
        <div class="card">
          <div style="font-weight:700; font-size:1.05rem;">{title}</div>
          <div style="opacity:0.85; margin-top:6px; line-height:1.6;">{text}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


# =========================
# EXCEL STORE (robust)
# =========================
class ExcelStore:
    def __init__(self, path: str):
        self.path = path
        self.ensure_workbook()

    def _format_sheet(self, ws):
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["A"].width = 12
        # widen notes if exists
        for i, cell in enumerate(ws[1], start=1):
            if str(cell.value).lower() == "notes":
                ws.column_dimensions[get_column_letter(i)].width = 40

    def _create_new_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = DAILY_SHEET
        ws.append(DAILY_HEADERS)
        self._format_sheet(ws)

        ws2 = wb.create_sheet(PROJECTS_SHEET)
        ws2.append(PROJECT_HEADERS)
        self._format_sheet(ws2)

        ws3 = wb.create_sheet(MONTHLY_SHEET)
        ws3.append(MONTHLY_HEADERS)
        self._format_sheet(ws3)

        wb.save(self.path)

    def ensure_workbook(self):
        if not os.path.exists(self.path):
            self._create_new_workbook()
            return

        try:
            wb = load_workbook(self.path)
        except Exception:
            # backup corrupted file
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = self.path.replace(".xlsx", f"_backup_{ts}.xlsx")
            try:
                os.rename(self.path, backup)
            except Exception:
                pass
            self._create_new_workbook()
            return

        changed = False
        changed |= self._ensure_sheet(wb, DAILY_SHEET, DAILY_HEADERS)
        changed |= self._ensure_sheet(wb, PROJECTS_SHEET, PROJECT_HEADERS)
        changed |= self._ensure_sheet(wb, MONTHLY_SHEET, MONTHLY_HEADERS)
        if changed:
            wb.save(self.path)

    def _ensure_sheet(self, wb, name: str, headers: List[str]) -> bool:
        changed = False
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(headers)
            self._format_sheet(ws)
            return True

        ws = wb[name]
        if ws.max_row < 1:
            ws.append(headers)
            changed = True

        existing = [c.value for c in ws[1]]
        existing = [e if e is not None else "" for e in existing]

        for h in headers:
            if h not in existing:
                ws.cell(row=1, column=len(existing) + 1, value=h)
                existing.append(h)
                changed = True

        if changed:
            self._format_sheet(ws)
        return changed

    def _load(self):
        return load_workbook(self.path)

    def _headers_idx(self, ws) -> Dict[str, int]:
        headers = [c.value for c in ws[1]]
        return {str(h): i+1 for i, h in enumerate(headers) if h is not None}

    def upsert_by_key(self, sheet: str, key_col: str, key_val: str, data: Dict[str, Any]):
        wb = self._load()
        ws = wb[sheet]
        idx = self._headers_idx(ws)

        # find row
        target_row = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, idx[key_col]).value) == str(key_val):
                target_row = r
                break

        if target_row is None:
            target_row = ws.max_row + 1
            ws.cell(target_row, idx[key_col], key_val)

        for k, v in data.items():
            if k in idx:
                ws.cell(target_row, idx[k], v)

        wb.save(self.path)

    def append_row(self, sheet: str, data: Dict[str, Any]):
        wb = self._load()
        ws = wb[sheet]
        idx = self._headers_idx(ws)
        r = ws.max_row + 1
        for k, v in data.items():
            if k in idx:
                ws.cell(r, idx[k], v)
        wb.save(self.path)

    def read_df(self, sheet: str) -> pd.DataFrame:
        # pandas reads cleaner for analytics
        try:
            df = pd.read_excel(self.path, sheet_name=sheet, engine="openpyxl")
            return df
        except Exception:
            return pd.DataFrame()


# =========================
# DATA LOADERS
# =========================
@st.cache_data(show_spinner=False)
def load_all_data(path: str):
    store = ExcelStore(path)

    daily = store.read_df(DAILY_SHEET)
    projects = store.read_df(PROJECTS_SHEET)
    monthly = store.read_df(MONTHLY_SHEET)

    # daily cleanup
    if not daily.empty and "date" in daily.columns:
        daily["date"] = pd.to_datetime(daily["date"], errors="coerce").dt.date
    else:
        daily = pd.DataFrame(columns=DAILY_HEADERS)

    # numeric cols
    num_cols = [
        "income","spent","dsa_leetcode_solved","js_fe_min","js_be_min",
        "prayer_fajr","prayer_dhuhr","prayer_asr","prayer_maghrib","prayer_isha",
        "eng_listen_min","eng_read_min","eng_speak_min","eng_write_min",
        "gym_trained","calories","screen_time_min"
    ]
    for c in num_cols:
        if c in daily.columns:
            daily[c] = pd.to_numeric(daily[c], errors="coerce").fillna(0)

    # projects cleanup
    if projects.empty:
        projects = pd.DataFrame(columns=PROJECT_HEADERS)
    if "month" in projects.columns:
        projects["month"] = projects["month"].astype(str)

    # monthly cleanup
    if monthly.empty:
        monthly = pd.DataFrame(columns=MONTHLY_HEADERS)
    if "month" in monthly.columns:
        monthly["month"] = monthly["month"].astype(str)

    return daily, projects, monthly


def invalidate_cache():
    st.cache_data.clear()


# =========================
# ANALYTICS
# =========================
def month_kpis(daily: pd.DataFrame, projects: pd.DataFrame, monthly: pd.DataFrame, ym: str) -> Dict[str, Any]:
    y, m = parse_month(ym)
    start, end = month_date_range(y, m)
    dim = calendar.monthrange(y, m)[1]

    d = daily[(daily["date"] >= start) & (daily["date"] <= end)].copy() if not daily.empty else pd.DataFrame(columns=DAILY_HEADERS)

    income = float(d["income"].sum()) if "income" in d.columns else 0.0
    spent = float(d["spent"].sum()) if "spent" in d.columns else 0.0
    net = income - spent

    leet = int(d["dsa_leetcode_solved"].sum()) if "dsa_leetcode_solved" in d.columns else 0
    js_fe = int(d["js_fe_min"].sum()) if "js_fe_min" in d.columns else 0
    js_be = int(d["js_be_min"].sum()) if "js_be_min" in d.columns else 0

    prayers_done = 0
    if not d.empty:
        prayers_done = int(d.apply(prayer_count_row, axis=1).sum())
    prayers_possible = 5 * dim
    prayer_rate = prayers_done / prayers_possible if prayers_possible else 0.0

    engL = int(d["eng_listen_min"].sum()) if "eng_listen_min" in d.columns else 0
    engR = int(d["eng_read_min"].sum()) if "eng_read_min" in d.columns else 0
    engS = int(d["eng_speak_min"].sum()) if "eng_speak_min" in d.columns else 0
    engW = int(d["eng_write_min"].sum()) if "eng_write_min" in d.columns else 0

    gym_days = int((d["gym_trained"] == 1).sum()) if "gym_trained" in d.columns else 0
    calories_avg = float(d["calories"].sum()) / dim if "calories" in d.columns and dim else 0.0
    screen_avg = float(d["screen_time_min"].sum()) / dim if "screen_time_min" in d.columns and dim else 0.0

    p = projects[projects["month"] == ym].copy() if not projects.empty else pd.DataFrame(columns=PROJECT_HEADERS)
    proj_count = int(len(p))

    mrow = monthly[monthly["month"] == ym].iloc[0].to_dict() if (not monthly.empty and (monthly["month"] == ym).any()) else {}

    return dict(
        ym=ym, dim=dim, rows_logged=int(len(d)),
        income=income, spent=spent, net=net,
        leet=leet, js_fe=js_fe, js_be=js_be,
        prayers_done=prayers_done, prayers_possible=prayers_possible, prayer_rate=prayer_rate,
        engL=engL, engR=engR, engS=engS, engW=engW,
        gym_days=gym_days, calories_avg=calories_avg, screen_avg=screen_avg,
        proj_count=proj_count, proj_df=p,
        monthly_row=mrow,
        daily_df=d
    )


def year_kpis(daily: pd.DataFrame, projects: pd.DataFrame, year: int) -> pd.DataFrame:
    rows = []
    for m in range(1, 13):
        ym = f"{year:04d}-{m:02d}"
        k = month_kpis(daily, projects, pd.DataFrame(), ym)
        rows.append({
            "month": ym,
            "income": k["income"],
            "spent": k["spent"],
            "net": k["net"],
            "leet": k["leet"],
            "js_min": k["js_fe"] + k["js_be"],
            "gym_days": k["gym_days"],
            "screen_avg": k["screen_avg"],
            "projects": k["proj_count"],
        })
    return pd.DataFrame(rows)


# =========================
# PAGES
# =========================
def page_daily_log(store: ExcelStore, daily: pd.DataFrame):
    st.header("Daily Log")
    st.caption("Log your daily numbers. This writes directly into Excel.")

    c1, c2 = st.columns([1.2, 1.0], gap="large")

    with c1:
        d = st.date_input("Date", value=today_date())
        ds = d.isoformat()

        # load existing row if exists
        row = {}
        if not daily.empty and (daily["date"] == d).any():
            row = daily[daily["date"] == d].iloc[0].to_dict()

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        st.subheader("Money")
        m1, m2, m3 = st.columns(3)
        income = m1.number_input("Income (day)", min_value=0.0, value=float(row.get("income", 0.0) or 0.0), step=10.0)
        spent = m2.number_input("Spent (day)", min_value=0.0, value=float(row.get("spent", 0.0) or 0.0), step=10.0)
        m3.metric("Net (day)", f"{income - spent:,.2f}")

        st.subheader("Skills (Daily)")
        s1, s2, s3 = st.columns(3)
        leet = s1.number_input("DSA LeetCode solved", min_value=0, value=int(row.get("dsa_leetcode_solved", 0) or 0), step=1)
        js_fe = s2.number_input("JS Frontend minutes", min_value=0, value=int(row.get("js_fe_min", 0) or 0), step=10)
        js_be = s3.number_input("JS Backend minutes", min_value=0, value=int(row.get("js_be_min", 0) or 0), step=10)

        st.subheader("Prayer")
        pcols = st.columns(5)
        pf = pcols[0].checkbox("Fajr", value=bool(int(row.get("prayer_fajr", 0) or 0)))
        pdh = pcols[1].checkbox("Dhuhr", value=bool(int(row.get("prayer_dhuhr", 0) or 0)))
        pasr = pcols[2].checkbox("Asr", value=bool(int(row.get("prayer_asr", 0) or 0)))
        pm = pcols[3].checkbox("Maghrib", value=bool(int(row.get("prayer_maghrib", 0) or 0)))
        pi = pcols[4].checkbox("Isha", value=bool(int(row.get("prayer_isha", 0) or 0)))

        st.subheader("English (minutes)")
        e1, e2, e3, e4 = st.columns(4)
        engL = e1.number_input("Listening", min_value=0, value=int(row.get("eng_listen_min", 0) or 0), step=10)
        engR = e2.number_input("Reading", min_value=0, value=int(row.get("eng_read_min", 0) or 0), step=10)
        engS = e3.number_input("Speaking", min_value=0, value=int(row.get("eng_speak_min", 0) or 0), step=10)
        engW = e4.number_input("Writing", min_value=0, value=int(row.get("eng_write_min", 0) or 0), step=10)

        st.subheader("Gym & Health (Daily)")
        g1, g2, g3, g4 = st.columns([1.1, 1, 1, 1.1])
        gym = g1.checkbox("Trained today", value=bool(int(row.get("gym_trained", 0) or 0)))
        calories = g2.number_input("Calories", min_value=0, value=int(row.get("calories", 0) or 0), step=50)
        screen = g3.number_input("Screen time (min)", min_value=0, value=int(row.get("screen_time_min", 0) or 0), step=15)
        g4.metric("Total practice (min)", f"{js_fe + js_be + engL + engR + engS + engW:,}")

        notes = st.text_area("Notes", value=str(row.get("notes", "") or ""), height=120)

        if st.button("Save Daily Log ✅"):
            data = {
                "date": ds,
                "income": float(income),
                "spent": float(spent),
                "dsa_leetcode_solved": int(leet),
                "js_fe_min": int(js_fe),
                "js_be_min": int(js_be),
                "prayer_fajr": 1 if pf else 0,
                "prayer_dhuhr": 1 if pdh else 0,
                "prayer_asr": 1 if pasr else 0,
                "prayer_maghrib": 1 if pm else 0,
                "prayer_isha": 1 if pi else 0,
                "eng_listen_min": int(engL),
                "eng_read_min": int(engR),
                "eng_speak_min": int(engS),
                "eng_write_min": int(engW),
                "gym_trained": 1 if gym else 0,
                "calories": int(calories),
                "screen_time_min": int(screen),
                "notes": notes.strip(),
            }
            try:
                store.upsert_by_key(DAILY_SHEET, "date", ds, data)
                invalidate_cache()
                st.success("Saved! (Excel updated)")
                st.rerun()
            except PermissionError:
                st.error("Excel file is open. Close life_tracker.xlsx in Excel and try again.")

    with c2:
        st.subheader("Quick Today Summary")
        card("Net (today)", f"{income - spent:,.2f}", "Income - Spent")
        card("LeetCode solved", f"{leet}", "Daily DSA output")
        card("JS minutes", f"{js_fe + js_be}", f"FE: {js_fe} | BE: {js_be}")
        card("Prayer", f"{(1 if pf else 0)+(1 if pdh else 0)+(1 if pasr else 0)+(1 if pm else 0)+(1 if pi else 0)}/5", "Check your consistency")
        card("English minutes", f"{engL + engR + engS + engW}", f"L:{engL} R:{engR} S:{engS} W:{engW}")
        card("Gym", "Yes ✅" if gym else "No", f"Calories: {calories} | Screen: {screen} min")

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
        info_box(
            "Pro tip",
            "If you want smarter insights: set monthly goals (LeetCode, JS minutes, English minutes, gym days). "
            "Then the dashboards can show goal progress automatically."
        )


def page_monthly_levels(store: ExcelStore, monthly: pd.DataFrame):
    st.header("Monthly Levels")
    st.caption("Track your monthly levels (English + Body fat + optional self-levels).")

    ym = st.text_input("Month (YYYY-MM)", value=month_str(today_date()))

    # load existing
    row = {}
    if not monthly.empty and (monthly["month"] == ym).any():
        row = monthly[monthly["month"] == ym].iloc[0].to_dict()

    c1, c2 = st.columns([1.2, 1], gap="large")

    with c1:
        st.subheader("English Levels (0-10 or A1..C2)")
        e1, e2, e3, e4 = st.columns(4)
        engL = e1.text_input("Listening level", value=str(row.get("eng_listen_level", "") or ""))
        engR = e2.text_input("Reading level", value=str(row.get("eng_read_level", "") or ""))
        engS = e3.text_input("Speaking level", value=str(row.get("eng_speak_level", "") or ""))
        engW = e4.text_input("Writing level", value=str(row.get("eng_write_level", "") or ""))

        st.subheader("Optional self-levels")
        s1, s2, s3 = st.columns(3)
        dsa_level = s1.text_input("DSA level", value=str(row.get("dsa_level", "") or ""))
        js_fe_level = s2.text_input("JS FE level", value=str(row.get("js_fe_level", "") or ""))
        js_be_level = s3.text_input("JS BE level", value=str(row.get("js_be_level", "") or ""))

        st.subheader("Body Fat (monthly)")
        bf = st.text_input("Body fat %", value=str(row.get("body_fat_pct", "") or ""))

        notes = st.text_area("Monthly notes", value=str(row.get("notes", "") or ""), height=140)

        if st.button("Save Monthly Levels ✅"):
            bf_val = bf.strip()
            bf_num = None
            if bf_val != "":
                try:
                    bf_num = float(bf_val)
                except Exception:
                    st.error("Body fat must be a number like 18.5")
                    st.stop()

            data = {
                "month": ym,
                "eng_listen_level": engL.strip() or None,
                "eng_read_level": engR.strip() or None,
                "eng_speak_level": engS.strip() or None,
                "eng_write_level": engW.strip() or None,
                "dsa_level": dsa_level.strip() or None,
                "js_fe_level": js_fe_level.strip() or None,
                "js_be_level": js_be_level.strip() or None,
                "body_fat_pct": bf_num,
                "notes": notes.strip(),
            }
            try:
                store.upsert_by_key(MONTHLY_SHEET, "month", ym, data)
                invalidate_cache()
                st.success("Saved! (Excel updated)")
                st.rerun()
            except PermissionError:
                st.error("Excel file is open. Close life_tracker.xlsx in Excel and try again.")

    with c2:
        st.subheader("What should a 'level' mean?")
        info_box(
            "Recommendation",
            "For English: use a consistent scale (A1/A2/B1/B2/C1/C2) OR numeric 1–10.\n\n"
            "For DSA & JS: also keep a consistent scale, so your yearly chart makes sense."
        )


def page_projects(store: ExcelStore, projects: pd.DataFrame):
    st.header("Projects")
    st.caption("Log projects per month with track + difficulty/level.")

    left, right = st.columns([1.1, 1.3], gap="large")

    with left:
        d = st.date_input("Project log date", value=today_date(), key="proj_date")
        ym = month_str(d)

        track = st.selectbox("Track", ["DSA", "JS-FE", "JS-BE", "Other"])
        name = st.text_input("Project name")
        level = st.select_slider("Project level (1 easy → 5 hard)", options=[1, 2, 3, 4, 5], value=3)
        hours = st.number_input("Hours spent", min_value=0.0, value=0.0, step=1.0)
        link = st.text_input("Link (GitHub / demo)")
        notes = st.text_area("Notes", height=100)

        if st.button("Add Project ✅"):
            if not name.strip():
                st.error("Project name is required.")
                st.stop()

            data = {
                "date": d.isoformat(),
                "month": ym,
                "track": track,
                "project_name": name.strip(),
                "project_level": int(level),
                "hours_spent": float(hours),
                "link": link.strip(),
                "notes": notes.strip(),
            }
            try:
                store.append_row(PROJECTS_SHEET, data)
                invalidate_cache()
                st.success("Project added!")
                st.rerun()
            except PermissionError:
                st.error("Excel file is open. Close life_tracker.xlsx in Excel and try again.")

    with right:
        st.subheader("Browse Projects")
        filt_month = st.text_input("Filter month (YYYY-MM)", value=month_str(today_date()))
        filt_track = st.multiselect("Track filter", ["DSA", "JS-FE", "JS-BE", "Other"], default=["DSA", "JS-FE", "JS-BE", "Other"])

        df = projects.copy()
        if not df.empty:
            if "month" in df.columns:
                df = df[df["month"].astype(str) == filt_month]
            if "track" in df.columns:
                df = df[df["track"].isin(filt_track)]

        if df.empty:
            info_box("No projects found", "Add your first project on the left.")
        else:
            st.dataframe(df.sort_values(by=["date"], ascending=False), use_container_width=True, height=360)

            # summary chart
            if "project_level" in df.columns:
                lvl = df["project_level"].fillna(0).astype(int)
                chart_df = lvl.value_counts().sort_index().reset_index()
                chart_df.columns = ["level", "count"]
                fig = px.bar(chart_df, x="level", y="count", title="Projects by Level")
                st.plotly_chart(ensure_plotly_dark(fig), use_container_width=True)


def page_dashboard_day(daily: pd.DataFrame):
    st.header("Dashboard — Day")
    d = st.date_input("Pick a day", value=today_date(), key="dash_day")
    if daily.empty:
        info_box("No data yet", "Start by adding a Daily Log.")
        return

    row = daily[daily["date"] == d]
    if row.empty:
        info_box("No log for this day", "Go to Daily Log and save your data for this date.")
        return

    r = row.iloc[0]
    income = float(r["income"])
    spent = float(r["spent"])
    net = income - spent

    leet = int(r["dsa_leetcode_solved"])
    js = int(r["js_fe_min"] + r["js_be_min"])
    prayers = int(prayer_count_row(r))
    eng = int(r["eng_listen_min"] + r["eng_read_min"] + r["eng_speak_min"] + r["eng_write_min"])
    gym = int(r["gym_trained"])
    calories = int(r["calories"])
    screen = int(r["screen_time_min"])

    c = st.columns(6, gap="small")
    with c[0]: card("Net", f"{net:,.2f}")
    with c[1]: card("Income", f"{income:,.2f}")
    with c[2]: card("Spent", f"{spent:,.2f}")
    with c[3]: card("LeetCode", f"{leet}")
    with c[4]: card("English (min)", f"{eng}")
    with c[5]: card("Prayer", f"{prayers}/5")

    c2 = st.columns(4, gap="small")
    with c2[0]: card("JS (min)", f"{js}")
    with c2[1]: card("Gym", "Yes ✅" if gym == 1 else "No")
    with c2[2]: card("Calories", f"{calories}")
    with c2[3]: card("Screen (min)", f"{screen}")

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    # show last 14 days trend
    recent = daily.copy()
    recent = recent.dropna(subset=["date"])
    recent = recent.sort_values("date")
    recent = recent[recent["date"] <= d].tail(14)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=recent["date"], y=recent["income"], name="Income"))
    fig.add_trace(go.Scatter(x=recent["date"], y=recent["spent"], name="Spent"))
    fig.update_layout(title="Last 14 days — Income vs Spent")
    st.plotly_chart(ensure_plotly_dark(fig), use_container_width=True)

    fig2 = go.Figure()
    fig2.add_trace(go.Bar(x=recent["date"], y=recent["dsa_leetcode_solved"], name="LeetCode"))
    fig2.update_layout(title="Last 14 days — LeetCode solved")
    st.plotly_chart(ensure_plotly_dark(fig2), use_container_width=True)

    notes = str(r.get("notes", "") or "").strip()
    if notes:
        info_box("Notes", notes)


def page_dashboard_month(daily: pd.DataFrame, projects: pd.DataFrame, monthly: pd.DataFrame):
    st.header("Dashboard — Month")
    ym = st.text_input("Month (YYYY-MM)", value=month_str(today_date()), key="dash_month")

    try:
        k = month_kpis(daily, projects, monthly, ym)
    except Exception:
        st.error("Bad month format. Use YYYY-MM")
        return

    c = st.columns(6, gap="small")
    with c[0]: card("Net", f"{k['net']:,.2f}", "Income - Spent")
    with c[1]: card("Income", f"{k['income']:,.2f}")
    with c[2]: card("Spent", f"{k['spent']:,.2f}", f"Avg/day: {k['spent']/k['dim'] if k['dim'] else 0:,.2f}")
    with c[3]: card("LeetCode", f"{k['leet']}", f"Avg/day: {k['leet']/k['dim'] if k['dim'] else 0:.2f}")
    with c[4]: card("JS (min)", f"{k['js_fe']+k['js_be']}", f"FE: {k['js_fe']} | BE: {k['js_be']}")
    with c[5]: card("Prayer", f"{k['prayer_rate']*100:.1f}%", f"{k['prayers_done']}/{k['prayers_possible']}")

    c2 = st.columns(5, gap="small")
    with c2[0]: card("English (min)", f"{k['engL']+k['engR']+k['engS']+k['engW']}", f"L:{k['engL']} R:{k['engR']} S:{k['engS']} W:{k['engW']}")
    with c2[1]: card("Gym days", f"{k['gym_days']}", f"Out of {k['dim']}")
    with c2[2]: card("Avg calories/day", f"{k['calories_avg']:.0f}")
    with c2[3]: card("Avg screen/day", f"{k['screen_avg']:.0f} min")
    with c2[4]: card("Projects", f"{k['proj_count']}")

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    # Monthly levels panel
    mr = k["monthly_row"] or {}
    if mr:
        info_box(
            "Monthly Levels",
            f"English levels → L:{mr.get('eng_listen_level','')} R:{mr.get('eng_read_level','')} S:{mr.get('eng_speak_level','')} W:{mr.get('eng_write_level','')}\n\n"
            f"Body fat % → {mr.get('body_fat_pct','')}\n\n"
            f"DSA level → {mr.get('dsa_level','')} | JS FE → {mr.get('js_fe_level','')} | JS BE → {mr.get('js_be_level','')}"
        )
    else:
        info_box("Monthly Levels not set", "Go to Monthly Levels page and save levels for this month.")

    # Charts (daily within month)
    ddf = k["daily_df"]
    if ddf.empty:
        info_box("No daily logs this month", "Start logging daily entries to unlock charts.")
        return

    ddf = ddf.sort_values("date")
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=ddf["date"], y=ddf["income"], name="Income"))
    fig.add_trace(go.Scatter(x=ddf["date"], y=ddf["spent"], name="Spent"))
    fig.update_layout(title="Daily Income vs Spent")
    st.plotly_chart(ensure_plotly_dark(fig), use_container_width=True)

    fig2 = px.bar(ddf, x="date", y="dsa_leetcode_solved", title="LeetCode solved per day")
    st.plotly_chart(ensure_plotly_dark(fig2), use_container_width=True)

    # JS minutes split chart
    js_df = ddf[["date","js_fe_min","js_be_min"]].copy()
    js_df = js_df.rename(columns={"js_fe_min":"JS FE", "js_be_min":"JS BE"})
    js_long = js_df.melt(id_vars=["date"], var_name="track", value_name="minutes")
    fig3 = px.area(js_long, x="date", y="minutes", color="track", title="JS study minutes (FE vs BE)")
    st.plotly_chart(ensure_plotly_dark(fig3), use_container_width=True)

    # Projects by level pie / bar
    p = k["proj_df"]
    if not p.empty and "project_level" in p.columns:
        tmp = p["project_level"].fillna(0).astype(int).value_counts().sort_index().reset_index()
        tmp.columns = ["level","count"]
        fig4 = px.pie(tmp, names="level", values="count", title="Projects by level")
        st.plotly_chart(ensure_plotly_dark(fig4), use_container_width=True)


def page_dashboard_year(daily: pd.DataFrame, projects: pd.DataFrame, monthly: pd.DataFrame):
    st.header("Dashboard — Year")
    year = st.number_input("Year", min_value=2000, max_value=2100, value=today_date().year, step=1)

    ydf = year_kpis(daily, projects, int(year))
    c = st.columns(4, gap="small")

    with c[0]: card("Income (year)", f"{ydf['income'].sum():,.2f}")
    with c[1]: card("Spent (year)", f"{ydf['spent'].sum():,.2f}")
    with c[2]: card("Net (year)", f"{ydf['net'].sum():,.2f}")
    with c[3]: card("LeetCode (year)", f"{int(ydf['leet'].sum())}")

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=ydf["month"], y=ydf["net"], name="Net"))
    fig.update_layout(title="Net income by month")
    st.plotly_chart(ensure_plotly_dark(fig), use_container_width=True)

    fig2 = go.Figure()
    fig2.add_trace(go.Bar(x=ydf["month"], y=ydf["leet"], name="LeetCode"))
    fig2.update_layout(title="LeetCode solved by month")
    st.plotly_chart(ensure_plotly_dark(fig2), use_container_width=True)

    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=ydf["month"], y=ydf["gym_days"], name="Gym days"))
    fig3.update_layout(title="Gym days by month")
    st.plotly_chart(ensure_plotly_dark(fig3), use_container_width=True)

    # Body fat trend from monthly sheet (if exists)
    if not monthly.empty and "body_fat_pct" in monthly.columns:
        bf = monthly.copy()
        bf = bf[bf["month"].str.startswith(str(year))]
        if not bf.empty:
            bf["body_fat_pct"] = pd.to_numeric(bf["body_fat_pct"], errors="coerce")
            bf = bf.dropna(subset=["body_fat_pct"])
            if not bf.empty:
                fig4 = px.line(bf.sort_values("month"), x="month", y="body_fat_pct", title="Body fat % (monthly)")
                st.plotly_chart(ensure_plotly_dark(fig4), use_container_width=True)


# =========================
# MAIN
# =========================
def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")
    st.markdown(DARK_CSS, unsafe_allow_html=True)

    # Excel path in same folder
    base_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base_dir, EXCEL_FILE)

    store = ExcelStore(path)
    daily, projects, monthly = load_all_data(path)

    st.sidebar.title("📊 Life Tracker")
    st.sidebar.caption("Excel-backed • Dark UI • Clear dashboards")

    page = st.sidebar.radio(
        "Navigate",
        [
            "Daily Log",
            "Monthly Levels",
            "Projects",
            "Dashboard — Day",
            "Dashboard — Month",
            "Dashboard — Year",
        ],
        index=0
    )

    st.sidebar.markdown("---")
    st.sidebar.write("**Excel file:**")
    st.sidebar.code(path, language="text")
    st.sidebar.info("Close Excel while saving, otherwise Windows may block file writes.")

    if page == "Daily Log":
        page_daily_log(store, daily)
    elif page == "Monthly Levels":
        page_monthly_levels(store, monthly)
    elif page == "Projects":
        page_projects(store, projects)
    elif page == "Dashboard — Day":
        page_dashboard_day(daily)
    elif page == "Dashboard — Month":
        page_dashboard_month(daily, projects, monthly)
    elif page == "Dashboard — Year":
        page_dashboard_year(daily, projects, monthly)


if __name__ == "__main__":
    main()